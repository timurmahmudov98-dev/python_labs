import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys

sys.path.append(r"C:\Users\user\Desktop\python_labs\src")


def csv_to_xlsx(csv_path_1, xlsx_path_2):
    csv_newpath = Path(csv_path_1)
    xlsx_newpath = Path(xlsx_path_2)

    if not csv_newpath.exists():
        raise FileNotFoundError("Файла csv не существует")

    workbook = Workbook()  # создание книги эксель
    work_sheet = workbook.active  # активный лист
    work_sheet.title = "Sheet1"

    column_width = {}  # для макс ширины столбца

    with csv_newpath.open("r", encoding="utf-8") as cf:
        csv_reading = csv.reader(cf)
        for row in csv_reading:
            work_sheet.append(row)
            for column_ind, value in enumerate(
                row, start=1
            ):  # enumerate для получения (индекс, значение)
                cell_length = len(value)  # длина значения
                if cell_length > column_width.get(
                    column_ind, 0
                ):  # случай, если длина значения > текущей макс ширины
                    column_width[column_ind] = cell_length

    for column_ind, width in column_width.items():  # для индекса, ширины в словаре
        letter = get_column_letter(
            column_ind
        )  # исходя из индекса, присваеваем столбцам экселевские буквы
        work_sheet.column_dimensions[letter].width = (
            (width + 2) if (width + 2) >= 8 else 8
        )  # по 1 отступу с обеих сторон

    workbook.save(xlsx_newpath)
